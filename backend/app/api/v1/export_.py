import io
import logging
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from supabase import create_client, Client
from app.config import settings
from app.deps import verify_token
from typing import Optional
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

router = APIRouter()


def get_supabase():
    return create_client(settings.SUPABASE_URL, settings.SUPABASE_SERVICE_ROLE_KEY)


# ─── Shared helpers ───────────────────────────────────────────────────────────

def resolve_columns(
    requested: Optional[str], catalogue: list[tuple[str, str]]
) -> list[tuple[str, str]]:
    """Return catalogue entries matching the requested comma-separated keys, in canonical order."""
    if not requested:
        return catalogue
    keys = set(requested.split(","))
    return [(k, l) for k, l in catalogue if k in keys]


def format_mes_ano(val: str | None) -> str:
    if val and len(val) >= 7:
        parts = val.split("-")
        return f"{parts[1]}/{parts[0]}"
    return val or ""


PARCEIRO_LABELS = {
    "totalpass": "Totalpass",
    "wellhub": "Wellhub",
}


def format_parceiros(value) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return ", ".join(PARCEIRO_LABELS.get(str(v), str(v)) for v in value)
    return str(value)


def build_xlsx(title: str, active_columns: list[tuple[str, str]], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_, label) in enumerate(active_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_two_section_xlsx(
    active_columns: list[tuple[str, str]],
    section1_rows: list[list],
    section2_rows: list[list],
) -> bytes:
    """Build an XLSX with two labelled sections separated by a blank row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento Mensal"

    col_count = len(active_columns)

    def write_section_header(row_idx: int, label: str, fill_color: str):
        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_count > 1:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_count)

    def write_col_headers(row_idx: int, fill_color: str):
        font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, (_, label) in enumerate(active_columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.font = font
            cell.fill = fill
            cell.alignment = align

    def write_rows(start_row: int, rows: list[list]):
        for row_offset, row in enumerate(rows):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=start_row + row_offset, column=col_idx, value=value)

    # Section 1 — Faturamento Mensal (blue)
    write_section_header(1, "Faturamento Mensal", "2563EB")
    write_col_headers(2, "3B82F6")
    write_rows(3, section1_rows)

    # Blank separator
    blank_row = 3 + len(section1_rows)
    ws.row_dimensions[blank_row].height = 10

    # Section 2 — Subsídio (amber)
    sec2_start = blank_row + 1
    write_section_header(sec2_start, "Subsídio — Cartão Realizado < Contratado", "D97706")
    write_col_headers(sec2_start + 1, "F59E0B")
    write_rows(sec2_start + 2, section2_rows)

    # Auto-width based on all content
    for col_idx in range(1, col_count + 1):
        col_letter = ws.cell(row=2, column=col_idx).column_letter
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for c in cell:
                if isinstance(c, MergedCell):
                    continue
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def xlsx_response(content: bytes, filename: str) -> Response:
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Column catalogues ────────────────────────────────────────────────────────

MONTHLY_COLUMNS: list[tuple[str, str]] = [
    ("empresa", "Empresa"),
    ("parceiros", "Parceiros"),
    ("mes_ano", "Mês/Ano"),
    ("elegiveis_contrato", "Elegíveis Contrato"),
    ("elegiveis", "Elegíveis"),
    ("valor_elegivel", "Valor Elegível"),
    ("valor_final", "Valor Final"),
    ("vidas_cobradas", "Vidas Cobradas"),
    ("valor_vidas", "PRO RATA"),
    ("pro_rata_dependente", "PRO RATA Dependente"),
    ("qtd_dependentes_gympass", "Qtd de Dependentes"),
    ("custo_por_dependente", "Custo por Dependente"),
    ("total_custo_dependentes", "Total de Custo por Dependente"),
    ("nr_cartao_contrato_flex", "Nº Cartão Contrato Flex"),
    ("nr_cartao_carga_flex", "Nº Cartão Carga Flex"),
    ("rs_carregado", "R$ Carregado"),
    ("media_cartao_realizado", "Média Cartão Realizado"),
    ("media_contratada", "Média Contratada"),
    ("nr_vidas", "Nº Vidas"),
    ("valor_elegivel_wiipo", "Valor Elegível Wiipo"),
    ("faturamento_wiipo", "Faturamento Wiipo"),
    ("qtd_dependentes", "Qtd Dependentes"),
    ("valor_por_dependente", "Valor por Dependente"),
    ("mensal_x_rentabilidade", "Mensal x Rentabilidade"),
    ("custo_por_cliente", "Custo por Cliente"),
    ("faturamento", "Faturamento"),
    ("faturamento_dependentes", "Faturamento de Dependentes"),
    ("observacao", "Observação"),
]

# Company fields available in the rentabilidade export
COMPANY_COLUMNS: list[tuple[str, str]] = [
    ("company_id", "Company ID"),
    ("empresa", "Empresa"),
    ("cnpj", "CNPJ"),
    ("razao_social", "Razão Social"),
    ("data_assinatura_contrato", "Data Assinatura Contrato"),
    ("email_envio", "E-mail para Envio"),
    ("inicio_cobranca", "Início Cobrança"),
    ("vencimento", "Dia de Vencimento"),
    ("parceiros", "Parceiros"),
]

COMPANY_FIELD_KEYS = {k for k, _ in COMPANY_COLUMNS}

RENTABILIDADE_COLUMNS: list[tuple[str, str]] = COMPANY_COLUMNS + [
    ("mes_ano", "Mês/Ano"),
    ("elegiveis_contrato", "Elegíveis Contrato"),
    ("elegiveis", "Elegíveis"),
    ("valor_elegivel", "Valor Elegível"),
    ("valor_final", "Valor Final"),
    ("vidas_cobradas", "Vidas Cobradas"),
    ("valor_vidas", "PRO RATA"),
    ("pro_rata_dependente", "PRO RATA Dependente"),
    ("qtd_dependentes_gympass", "Qtd de Dependentes"),
    ("custo_por_dependente", "Custo por Dependente"),
    ("total_custo_dependentes", "Total de Custo por Dependente"),
    ("nr_cartao_contrato_flex", "Nº Cartão Contrato Flex"),
    ("nr_cartao_carga_flex", "Nº Cartão Carga Flex"),
    ("rs_carregado", "R$ Carregado"),
    ("media_cartao_realizado", "Média Cartão Realizado"),
    ("media_contratada", "Média Contratada"),
    ("nr_vidas", "Nº Vidas"),
    ("valor_elegivel_wiipo", "Valor Elegível Wiipo"),
    ("faturamento_wiipo", "Faturamento Wiipo"),
    ("qtd_dependentes", "Qtd Dependentes"),
    ("valor_por_dependente", "Valor por Dependente"),
    ("mensal_x_rentabilidade", "Mensal x Rentabilidade"),
    ("custo_por_cliente", "Custo por Cliente"),
    ("faturamento", "Faturamento"),
    ("faturamento_dependentes", "Faturamento de Dependentes"),
    ("observacao", "Observação"),
]


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/export/monthly")
def export_monthly_xlsx(
    mes_ano: Optional[str] = Query(None),
    columns: Optional[str] = Query(None, description="Comma-separated column keys"),
    _user=Depends(verify_token),
    supabase: Client = Depends(get_supabase),
):
    try:
        active_columns = resolve_columns(columns, MONTHLY_COLUMNS)

        records = (
            supabase.table("monthly_records").select("*")
            .eq("mes_ano", mes_ano) if mes_ano
            else supabase.table("monthly_records").select("*")
        ).execute().data or []

        company_map: dict[str, dict] = {}
        needs_company = any(k in ("empresa", "parceiros") for k, _ in active_columns)
        if needs_company and records:
            ids = list({r["company_id"] for r in records})
            res = (
                supabase.table("companies")
                .select("id, empresa, parceiros")
                .in_("id", ids)
                .execute()
            )
            company_map = {c["id"]: c for c in res.data or []}

        rows = []
        for rec in records:
            company = company_map.get(rec.get("company_id", ""), {})
            row = []
            for field, _ in active_columns:
                if field == "empresa":
                    row.append(company.get("empresa", ""))
                elif field == "parceiros":
                    row.append(format_parceiros(company.get("parceiros")))
                elif field == "mes_ano":
                    row.append(format_mes_ano(rec.get(field)))
                else:
                    row.append(rec.get(field))
            rows.append(row)

        content = build_xlsx("Registros Mensais", active_columns, rows)
        return xlsx_response(content, f"registros_mensais_{mes_ano or 'todos'}.xlsx")
    except Exception as e:
        logger.error("Error in export/monthly: %s", str(e), exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export monthly records: {str(e)}")


@router.get("/export/rentabilidade")
def export_rentabilidade_xlsx(
    mes_ano: str = Query(..., description="Month to export (YYYY-MM-DD)"),
    columns: Optional[str] = Query(None, description="Comma-separated column keys"),
    _user=Depends(verify_token),
    supabase: Client = Depends(get_supabase),
):
    try:
        active_columns = resolve_columns(columns, RENTABILIDADE_COLUMNS)

        # Section 1: records where mensal_x_rentabilidade = 'Faturamento mensal'
        sec1_result = (
            supabase.table("monthly_records")
            .select("*")
            .eq("mes_ano", mes_ano)
            .eq("mensal_x_rentabilidade", "Faturamento mensal")
            .execute()
        )
        sec1_records = sec1_result.data or []

        # Collect all company IDs needed across both sections
        sec1_company_ids = {r["company_id"] for r in sec1_records}

        # Section 2: companies with subsidio=true that have at least one record in this month
        # where media_cartao_realizado < media_contratada
        subsidio_result = (
            supabase.table("companies")
            .select("id, company_id, empresa, cnpj, razao_social, data_assinatura_contrato, email_envio, inicio_cobranca, vencimento, subsidio, parceiros")
            .eq("subsidio", True)
            .execute()
        )
        subsidio_companies = subsidio_result.data or []
        subsidio_company_ids = {c["id"] for c in subsidio_companies}

        sec2_records = []
        if subsidio_company_ids:
            all_records_for_month_result = (
                supabase.table("monthly_records")
                .select("*")
                .eq("mes_ano", mes_ano)
                .in_("company_id", list(subsidio_company_ids))
                .execute()
            )
            all_records_for_month = all_records_for_month_result.data or []
            for r in all_records_for_month:
                realizado = r.get("media_cartao_realizado")
                contratada = r.get("media_contratada")
                if realizado is not None and contratada is not None and realizado < contratada:
                    sec2_records.append(r)

        # Fetch company data for all needed IDs
        all_company_ids = sec1_company_ids | {r["company_id"] for r in sec2_records}
        company_map: dict[str, dict] = {}
        if all_company_ids:
            companies_res = (
                supabase.table("companies")
                .select("id, company_id, empresa, cnpj, razao_social, data_assinatura_contrato, email_envio, inicio_cobranca, vencimento, subsidio, parceiros")
                .in_("id", list(all_company_ids))
                .execute()
            )
            company_map = {c["id"]: c for c in companies_res.data or []}

        def build_rows(records: list) -> list[list]:
            rows = []
            for rec in records:
                company = company_map.get(rec.get("company_id", ""), {})
                row = []
                for field, _ in active_columns:
                    if field in COMPANY_FIELD_KEYS:
                        value = company.get(field)
                        if field == "parceiros":
                            value = format_parceiros(value)
                        elif field in ("data_assinatura_contrato", "inicio_cobranca") and value:
                            value = str(value)[:10].replace("-", "/")
                            parts = value.split("/")
                            if len(parts) == 3:
                                value = f"{parts[2]}/{parts[1]}/{parts[0]}"
                    elif field == "mes_ano":
                        value = format_mes_ano(rec.get(field))
                    else:
                        value = rec.get(field)
                    row.append(value)
                rows.append(row)
            return rows

        sec1_rows = build_rows(sec1_records)
        sec2_rows = build_rows(sec2_records)

        content = build_two_section_xlsx(active_columns, sec1_rows, sec2_rows)
        return xlsx_response(content, f"faturamento_mensal_{mes_ano}.xlsx")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error in export/rentabilidade: %s", str(e), exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export rentabilidade report: {str(e)}")


MONTHS_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def build_comparacao_mes_a_mes_xlsx(year: int, months: list[int], by_month: dict[str, dict]) -> bytes:
    """Matrix: metric rows × month columns (same KPI formulas as dashboard)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparação Mês a Mês"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_font = Font(bold=True)

    ws.cell(row=1, column=1, value="").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for col_idx, month in enumerate(months, start=2):
        cell = ws.cell(row=1, column=col_idx, value=f"{MONTHS_PT[month - 1]}/{year}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    metric_rows = [
        ("Empresas matriz", "empresas_ativas"),
        ("Empresas filiais", "empresas_filiais"),
        ("Total de empresas", "total_empresas"),
        ("total de vidas cobradas", "total_vidas_cobradas"),
        ("Total Custo por Vida", "total_valor_vidas"),
        ("total custo por cliente", "total_custo_por_cliente"),
        ("total de faturamento", "total_faturamento"),
    ]

    money_keys = {"total_valor_vidas", "total_custo_por_cliente", "total_faturamento"}

    for row_idx, (label, key) in enumerate(metric_rows, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = label_font
        for col_idx, month in enumerate(months, start=2):
            mes_key = f"{year}-{str(month).zfill(2)}-01"
            bucket = by_month.get(mes_key, {})
            value = bucket.get(key, 0)
            if key in money_keys:
                cell = ws.cell(row=row_idx, column=col_idx, value=float(value or 0))
                cell.number_format = 'R$ #,##0.00'
            else:
                ws.cell(row=row_idx, column=col_idx, value=int(value or 0))

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(months) + 2):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/export/comparacao-mes-a-mes")
def export_comparacao_mes_a_mes_xlsx(
    year: int = Query(..., description="Year to compare (YYYY)", ge=2020, le=2099),
    _user=Depends(verify_token),
    supabase: Client = Depends(get_supabase),
):
    """
    Month-to-month KPI matrix for the year.
    Same universe as the Dashboard: active companies with a monthly record in each month.
    """
    try:
        from datetime import date
        from app.api.v1.dashboard import (
            RECORD_SELECT,
            accumulate_kpi_record,
            empty_kpi_bucket,
            fetch_active_company_tipos,
            finalize_kpi_bucket,
        )

        today = date.today()
        if year > today.year:
            months: list[int] = []
        elif year < today.year:
            months = list(range(1, 13))
        else:
            months = list(range(1, today.month + 1))

        company_tipos = fetch_active_company_tipos(supabase)
        active_company_ids = list(company_tipos.keys())

        by_month: dict[str, dict] = {
            f"{year}-{str(m).zfill(2)}-01": empty_kpi_bucket()
            for m in months
        }

        if months and active_company_ids:
            start = f"{year}-{str(months[0]).zfill(2)}-01"
            end = f"{year}-{str(months[-1]).zfill(2)}-01"
            result = (
                supabase.table("monthly_records")
                .select(RECORD_SELECT)
                .in_("company_id", active_company_ids)
                .gte("mes_ano", start)
                .lte("mes_ano", end)
                .execute()
            )
            for rec in result.data or []:
                raw = rec.get("mes_ano") or ""
                key = raw[:10] if len(raw) >= 10 else raw
                if key not in by_month:
                    continue
                accumulate_kpi_record(by_month[key], rec)

        finalized = {
            key: finalize_kpi_bucket(bucket, company_tipos)
            for key, bucket in by_month.items()
        }

        content = build_comparacao_mes_a_mes_xlsx(year, months, finalized)
        return xlsx_response(content, f"comparacao_mes_a_mes_{year}.xlsx")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error in export/comparacao-mes-a-mes: %s", str(e), exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export comparação mês a mês: {str(e)}")
