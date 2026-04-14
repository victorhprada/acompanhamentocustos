import io
import re
import uuid
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from supabase import create_client, Client

from app.config import settings
from app.deps import verify_token

router = APIRouter()

HEADER_ROW = 8
BUCKET = "imports"

COMPANY_FIELDS = {
    "company_id", "empresa", "cnpj", "razao_social", "cliente",
    "email_envio", "inicio_cobranca", "vencimento", "nota_fiscal_descricao",
}

MONTHLY_FIELDS = {
    "elegiveis_contrato", "elegiveis", "valor_elegivel", "valor_final",
    "elegiveis_totalpass_gympass", "vidas_cobradas", "nr_vidas", "valor_vidas",
    "nr_cartao_contrato_flex", "nr_cartao_carga_flex", "rs_carregado",
    "media_cartao_realizado", "media_contratada", "valor_elegivel_wiipo",
    "faturamento_wiipo", "mensal_x_rentabilidade", "custo_por_cliente",
    "valor_faturado", "faturamento",
}

NUMERIC_MONTHLY_FIELDS = MONTHLY_FIELDS - {"mensal_x_rentabilidade"}

PT_MONTHS = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def get_supabase() -> Client:
    return create_client(settings.SUPABASE_URL, settings.SUPABASE_SERVICE_ROLE_KEY)


def _db_code(exc: Exception) -> Optional[str]:
    """Extract PostgreSQL error code from a Supabase/postgrest exception."""
    try:
        return exc.code  # type: ignore[attr-defined]
    except AttributeError:
        pass
    try:
        return exc.args[0].get("code")  # type: ignore[index]
    except (AttributeError, IndexError, TypeError):
        return None


def _friendly_error(exc: Exception, label: str) -> str:
    """Return a human-readable Portuguese error message for common DB errors."""
    code = _db_code(exc)
    raw = str(exc)

    if code == "23505":
        # Extract which key/value conflicted
        details = ""
        try:
            details = exc.args[0].get("details", "") or ""  # type: ignore[index]
        except (AttributeError, IndexError, TypeError):
            pass
        m = re.search(r"Key \((\w+)\)=\(([^)]+)\)", details)
        if m:
            field, value = m.group(1), m.group(2)
            hints = {
                "company_id": f"Company ID '{value}' já cadastrado. O sistema usará o CNPJ como identificador alternativo.",
                "cnpj": f"CNPJ '{value}' já cadastrado — empresa atualizada.",
            }
            return f"{label}: {hints.get(field, f'valor duplicado no campo {field}: {value}')}"
        return f"{label}: registro duplicado — verifique se a empresa já está cadastrada"

    if code == "23502":
        m = re.search(r'column "(\w+)"', raw)
        field = m.group(1) if m else "campo desconhecido"
        field_labels = {
            "empresa": "Empresa (nome fantasia)",
            "cnpj": "CNPJ",
            "company_id": "Company ID",
        }
        return f"{label}: campo obrigatório '{field_labels.get(field, field)}' está vazio na planilha — verifique o mapeamento de colunas"

    if code == "22001":
        return f"{label}: valor muito longo em algum campo — verifique se o CNPJ tem espaços ou caracteres extras"

    return f"{label}: {raw}"


def ensure_bucket(supabase: Client) -> None:
    try:
        buckets = supabase.storage.list_buckets()
        if any(b.name == BUCKET for b in buckets):
            return  # already exists
        supabase.storage.create_bucket(BUCKET, options={"public": False})
    except Exception:
        pass  # ignore — create_bucket will be tried on next request


def detect_month(sheet_name: str) -> Optional[str]:
    """Detect mes_ano (YYYY-MM-01) from sheet name. Supports PT month names and numeric formats."""
    name = sheet_name.lower().strip()

    # "2026-01" or "01/2026"
    m = re.search(r"(\d{4})[/-](\d{1,2})", name)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}-01"

    m = re.search(r"(\d{1,2})[/-](\d{4})", name)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}-01"

    # "Janeiro 2026", "jan/2026", "março2026"
    for pt_name, month_num in PT_MONTHS.items():
        if pt_name in name:
            year_m = re.search(r"(\d{4})", name)
            year = int(year_m.group(1)) if year_m else date.today().year
            return f"{year}-{month_num:02d}-01"

    return None


def parse_currency(value) -> Optional[float]:
    """Parse Brazilian currency string to float. R$ 3.048,89 → 3048.89"""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("-", "", "None", "nan"):
        return None
    s = re.sub(r"[R$\s]", "", s)
    # Brazilian format: 3.048,89 → remove thousands dot, swap decimal comma
    if re.search(r"\d\.\d{3},\d", s):
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_date_br(value) -> Optional[str]:
    """Parse dd/mm/yyyy (or datetime objects) to YYYY-MM-DD."""
    import datetime as _dt
    if value is None:
        return None
    # openpyxl returns datetime objects for date-formatted cells
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if s in ("", "None", "nan", "-"):
        return None
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    # already ISO — return only the date portion (strip time if present)
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def parse_cell(value, system_field: str):
    """Convert raw cell value to the appropriate Python type for system_field."""
    if value is None:
        return None
    raw = str(value).strip()
    if raw in ("", "None", "nan", "-"):
        return None

    if system_field in NUMERIC_MONTHLY_FIELDS:
        return parse_currency(value)
    if system_field == "inicio_cobranca":
        return parse_date_br(value)
    if system_field == "vencimento":
        # Full date like "25/03/2026" or "2026-03-25" → extract day only
        m = re.match(r"(\d{1,2})[/-]\d{1,2}[/-]\d{2,4}", raw)
        if m:
            return int(m.group(1))
        m = re.match(r"\d{4}[/-]\d{1,2}[/-](\d{1,2})", raw)
        if m:
            return int(m.group(1))
        try:
            return int(float(raw))
        except (ValueError, TypeError):
            return None
    return raw


def extract_sheet_meta(ws) -> dict:
    """Extract column names and preview rows from a worksheet."""
    header = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
    columns = []
    for i, c in enumerate(header):
        label = str(c).strip() if c is not None else None
        if label and label not in ("None", "nan"):
            columns.append({"index": i, "label": label})

    preview = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=HEADER_ROW + 3, values_only=True):
        if any(cell is not None for cell in row):
            preview.append([str(c) if c is not None else "" for c in row])

    return {"columns": columns, "preview": preview}


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/import/upload")
def upload_import_file(
    file: UploadFile = File(...),
    _user=Depends(verify_token),
    supabase: Client = Depends(get_supabase),
):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Apenas arquivos Excel (.xlsx, .xls) são aceitos")

    content = file.file.read()
    file_path = f"{uuid.uuid4()}/{file.filename}"

    ensure_bucket(supabase)
    supabase.storage.from_(BUCKET).upload(
        file_path,
        content,
        {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        meta = extract_sheet_meta(ws)
        sheets.append({
            "name": sheet_name,
            "mes_ano": detect_month(sheet_name),
            "columns": meta["columns"],
            "preview": meta["preview"],
        })

    return {"file_path": file_path, "sheets": sheets}


MONTHS_2026 = [f"2026-{m:02d}-01" for m in range(1, 13)]


def _propagate_records(supabase: Client, company_db_id: str, mes_ano: str, monthly_data: dict, results: dict) -> None:
    """Replicate monthly_data to all future months in 2026 after mes_ano."""
    future_months = [m for m in MONTHS_2026 if m > mes_ano]
    for future_month in future_months:
        payload = {**monthly_data, "mes_ano": future_month}
        try:
            check = (
                supabase.table("monthly_records")
                .select("id")
                .eq("company_id", company_db_id)
                .eq("mes_ano", future_month)
                .execute()
            )
            if check.data:
                supabase.table("monthly_records").update(payload).eq("id", check.data[0]["id"]).execute()
            else:
                supabase.table("monthly_records").insert(payload).execute()
        except Exception:
            pass  # propagation errors are silent — base record already saved


@router.post("/import/process")
def process_import(
    body: dict,
    _user=Depends(verify_token),
    supabase: Client = Depends(get_supabase),
):
    file_path: str = body.get("file_path", "")
    mapping: dict = body.get("mapping", {})        # excel_label → system_field | "_skip"
    sheets_cfg: list = body.get("sheets", [])       # [{ name, mes_ano, include }]
    propagate: bool = body.get("propagate", False)
    propagate_mes_ano: Optional[str] = body.get("propagate_mes_ano")

    if not file_path:
        raise HTTPException(status_code=400, detail="file_path é obrigatório")
    if not mapping:
        raise HTTPException(status_code=400, detail="mapping é obrigatório")

    file_bytes = supabase.storage.from_(BUCKET).download(file_path)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    results = {
        "companies_created": 0,
        "companies_updated": 0,
        "records_created": 0,
        "records_updated": 0,
        "errors": [],
    }

    included_sheets = {s["name"]: s for s in sheets_cfg if s.get("include")}

    for sheet_name, sheet_cfg in included_sheets.items():
        if sheet_name not in wb.sheetnames:
            results["errors"].append(f"Aba '{sheet_name}' não encontrada no arquivo")
            continue

        mes_ano: Optional[str] = sheet_cfg.get("mes_ano")
        if not mes_ano:
            results["errors"].append(f"Aba '{sheet_name}': mês não definido — ignorada")
            continue

        ws = wb[sheet_name]
        raw_header = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
        # col_index → system_field (only mapped, non-skip fields)
        col_map: dict[int, str] = {}
        for i, cell in enumerate(raw_header):
            label = str(cell).strip() if cell is not None else None
            if label and label in mapping and mapping[label] not in ("_skip", "", None):
                col_map[i] = mapping[label]

        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            if not any(c is not None for c in row):
                continue

            row_data: dict = {}
            for i, cell in enumerate(row):
                if i in col_map:
                    system_field = col_map[i]
                    parsed = parse_cell(cell, system_field)
                    if parsed is not None:
                        row_data[system_field] = parsed

            # Skip empty / header-repeat rows
            if not row_data.get("cnpj") and not row_data.get("empresa"):
                continue

            company_data = {k: v for k, v in row_data.items() if k in COMPANY_FIELDS}
            monthly_data = {k: v for k, v in row_data.items() if k in MONTHLY_FIELDS}

            # --- Sanitize CNPJ ---
            if company_data.get("cnpj"):
                cnpj_raw = str(company_data["cnpj"]).strip()
                # Excel strips leading zeros from numeric CNPJs → pad to 14 digits
                digits_only = re.sub(r"\D", "", cnpj_raw)
                if digits_only == cnpj_raw and len(digits_only) < 14:
                    cnpj_raw = digits_only.zfill(14)
                company_data["cnpj"] = cnpj_raw[:18]

            # --- empresa fallback: use razao_social or cliente if empresa is missing ---
            if not company_data.get("empresa"):
                company_data["empresa"] = company_data.get("razao_social") or company_data.get("cliente")

            # --- Upsert company ---
            company_db_id: Optional[str] = None
            cnpj = company_data.get("cnpj")
            label = company_data.get("empresa") or cnpj or "?"

            if not cnpj:
                results["errors"].append(f"Linha ignorada: CNPJ ausente para '{label}' — mapeie a coluna CNPJ")
                continue

            if not company_data.get("empresa"):
                results["errors"].append(
                    f"Linha ignorada: campo 'Empresa' vazio para CNPJ {cnpj} — mapeie a coluna Empresa ou Razão Social"
                )
                continue

            try:
                existing = supabase.table("companies").select("id").eq("cnpj", cnpj).execute()
                if existing.data:
                    company_db_id = existing.data[0]["id"]
                    supabase.table("companies").update(company_data).eq("id", company_db_id).execute()
                    results["companies_updated"] += 1
                else:
                    # Ensure unique company_id; default to CNPJ digits
                    if not company_data.get("company_id"):
                        company_data["company_id"] = re.sub(r"\D", "", cnpj)[:50]
                    try:
                        ins = supabase.table("companies").insert(company_data).execute()
                        if ins.data:
                            company_db_id = ins.data[0]["id"]
                            results["companies_created"] += 1
                    except Exception as insert_exc:
                        if _db_code(insert_exc) == "23505" and "company_id" in str(insert_exc):
                            # company_id conflict → fall back to CNPJ digits unconditionally
                            company_data["company_id"] = re.sub(r"\D", "", cnpj)[:50]
                            ins2 = supabase.table("companies").insert(company_data).execute()
                            if ins2.data:
                                company_db_id = ins2.data[0]["id"]
                                results["companies_created"] += 1
                        else:
                            results["errors"].append(_friendly_error(insert_exc, f"Empresa '{label}'"))
                            continue
            except Exception as exc:
                results["errors"].append(_friendly_error(exc, f"Empresa '{label}'"))
                continue

            # --- Upsert monthly record ---
            if not company_db_id:
                continue

            monthly_data["company_id"] = company_db_id
            monthly_data["mes_ano"] = mes_ano

            try:
                existing_rec = (
                    supabase.table("monthly_records")
                    .select("id")
                    .eq("company_id", company_db_id)
                    .eq("mes_ano", mes_ano)
                    .execute()
                )
                if existing_rec.data:
                    supabase.table("monthly_records").update(monthly_data).eq("id", existing_rec.data[0]["id"]).execute()
                    results["records_updated"] += 1
                else:
                    supabase.table("monthly_records").insert(monthly_data).execute()
                    results["records_created"] += 1

                if propagate and propagate_mes_ano and mes_ano == propagate_mes_ano:
                    _propagate_records(supabase, company_db_id, mes_ano, monthly_data, results)
            except Exception as exc:
                results["errors"].append(
                    _friendly_error(exc, f"Registro {mes_ano} / {label}")
                )

    return results
